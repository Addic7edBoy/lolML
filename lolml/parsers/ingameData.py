# Import libraries
import requests
from bs4 import BeautifulSoup
import json
import pandas as pd
import glob
import datetime
import xlrd
import openpyxl as xl
import re


def convert_excel_time(excel_time):
    return pd.to_datetime('1899-12-30') + pd.to_timedelta(excel_time, 'D')


def save_xlsx(year='2019', d_split_arr=['spring', 'summer', 'worlds', 'complete'], path='../temp/inGame/'):
    for split in d_split_arr:
        resp = get_xlsx_request('https://oracleselixir.com/gamedata/' +
                                str(year) + '-' + split + '/', str(year), split)
        if str(resp) == '<Response [404]>':
            print('url not found')
            pass
        else:
            with open(path + year + '_' + split + '.xlsx', 'wb') as output:
                output.write(resp.content)


def get_xlsx_request(url, year, split):

    url = url
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'en-US,en;q=0.9,ru;q=0.8',
        'Connection': 'keep-alive',
        'Cookie': 'ga=GA1.2.241321163.1584212733; kola.128265=F35EF6A8-B1C7-4242-A989-2DB6A5687A65; kola.128265.session=F060AD5B-797E-4FC2-80C3-0506652403E8; _gid=GA1.2.510345096.1585323676,',
        'Host': 'oracleselixir.com',
        'Referer': 'https://oracleselixir.com/match-data/',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/80.0.3987.87 Chrome/80.0.3987.87 Safari/537.36',
    }

    resp = requests.get(url, headers=headers)
    return resp


def concat_tables(inGameDataPath='../temp/inGame/'):

    # filenames
    excel_names = glob.glob(inGameDataPath + "20*.xlsx")
    print(excel_names)

    # read them in
    # excels = [pd.ExcelFile(name) for name in excel_names]

    frames = [pd.read_excel(name) for name in excel_names]

    # turn them into dataframes
    # frames = [x.parse(x.sheet_names[0], header=None, index_col=None)
              # for x in excels]

    # drop all rows except team info ones
    frames = [teamOnly(df) for df in frames]

    # drop unnecessary columns
    frames = [drop(df) for df in frames]
    print(frames[0].head(10))

    # delete the first row for all frames except the first
    # i.e. remove the header row -- assumes it's the first
    frames[1:] = [df[1:] for df in frames[1:]]

    # concatenate them..
    combined = pd.concat(frames)

    # write it out
    combined.to_csv(inGameDataPath + "combined.csv", header=True, index=True)
    # combined.to_excel("combined.xlsx", header=False, index=False)


def xldate_to_datetime(xldate):
    temp = datetime.datetime(1899, 12, 30)
    delta = datetime.timedelta(days=xldate)
    return temp + delta


def teamOnly(dataObj):
    # table_raw = pd.read_csv(path)
    table_raw = dataObj
    print(type(table_raw))
    team_rows = table_raw[table_raw['playerid'].isin([100, 200])]
    return team_rows


def drop(dataObj):
    df = dataObj
    df1 = df[['league', 'split', 'week', 'game', 'date', 'patchno', 'side', 'team',
              'teamdragkills', 'elders', 'herald', 'teambaronkills', 'totalgold', 'earnedgpm', 'cspm']]
    df1 = df1[~(df1.league == 'LPL')]
    df1 = df1[~(df1.date == ' ')]
    return df1


def format_date(path):
    df = pd.read_csv(path)
    # df.date = df.date.map(lambda a: re.split(r'\.', a)[0])
    print(df.date)
    # df['date'] = df['date'].map(lambda a: xldate_to_datetime(int(re.split(r'\.', a.strip())[0])))
    df.date = (pd.to_datetime('1899-12-30') + pd.to_timedelta(df.date, 'D'))
    print(df.date)
    df.to_csv('../temp/inGame/result.csv', header=True, index=False)


def main():
    save_xlsx()
    concat_tables()
    format_date('../temp/inGame/combined.csv')


# main()

df = pd.read_csv('../temp/inGame/result.csv')
print(df.league.unique())
# print(type(df['date']))
# # print(df.at[12863, 'date'])
# print(df.loc[3213,'date'])
# print(type(df.loc[3213,'date']))
# if df.loc[3213,'date'] == '':
#     print('empty')
# elif df.loc[3213,'date'] == ' ':
#     print('spaces')
# elif df.loc[3213,'date'] == None:
#     print('None')
# else:
#     print('hui znaet')
# rows = drop('./spring.csv')
# rows.to_csv("c.csv", header=True, index=True)



"""
wb = xl.load_workbook('c.xlsx', enumerate)
sheet = wb.worksheets[0]

row_count = sheet.max_row
column_count = sheet.max_column
print(row_count, column_count)

print(result)

"""
